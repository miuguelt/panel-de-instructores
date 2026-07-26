import unittest
from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.utils.datetime import to_excel
from werkzeug.datastructures import FileStorage

from app import create_app, db
from app.models import (
    Aprendiz,
    ConfiguracionAlertas,
    ConfiguracionAseo,
    ConfiguracionRanking,
    Ficha,
    FichaInstructor,
    Instructor,
    JuicioEvaluativo,
    JuicioEvaluativoInstructor,
)
from app.services.importacion_ficha import ErrorImportacion, importar_archivo


class ImportacionFichaTestCase(unittest.TestCase):
    def setUp(self):
        self.app = create_app({
            'TESTING': True,
            'SQLALCHEMY_DATABASE_URI': 'sqlite:///:memory:',
            'SQLALCHEMY_ENGINE_OPTIONS': {},
            'WTF_CSRF_ENABLED': False,
            'RATELIMIT_ENABLED': False,
        })
        self.contexto = self.app.app_context()
        self.contexto.push()
        db.create_all()
        self.a = Instructor(nombre='Uno', correo='uno@sena.edu.co')
        self.a.set_password('x')
        self.b = Instructor(nombre='Dos', correo='dos@sena.edu.co')
        self.b.set_password('x')
        db.session.add_all([self.a, self.b])
        db.session.flush()
        self.ficha = Ficha(codigo='2672089', nombre_programa='Programa', instructor_id=self.a.id)
        db.session.add(self.ficha)
        db.session.commit()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.contexto.pop()

    def _archivo(self):
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.append(['Ficha de Caracterización:', '', 2672089])
        hoja.append(['Cógigo:', '', '228117'])
        hoja.append(['Denominación:', '', 'Programa'])
        for _ in range(10):
            hoja.append([])
        hoja.append(['Tipo de Documento', 'Número de Documento', 'Nombre', 'Apellidos', 'Estado',
                     'Competencia', 'Resultado de Aprendizaje', 'Juicio de Evaluación',
                     'Fecha y Hora del Juicio Evaluativo', 'Funcionario que registro el juicio evaluativo'])
        fila = ['CC', '1001', 'Ana', 'Pérez', 'EN_FORMACION', 'Competencia 1', 'Resultado 1',
                'APROBADO', '14/03/2023', 'Instructor anterior']
        hoja.append(fila)
        hoja.append(fila)
        salida = BytesIO()
        libro.save(salida)
        salida.seek(0)
        return FileStorage(stream=salida, filename='reporte.xlsx')

    def _reporte_para_crear(self):
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.append(['Reporte de Juicios de Evaluación'])
        hoja.append(['Fecha del Reporte:', '24/06/2026'])
        hoja.append(['Ficha de Caracterización:', 3336360])
        hoja.append(['Código:', 133100])
        hoja.append(['Versión:', 2])
        hoja.append(['Denominación:', 'CONTABILIZACIÓN DE OPERACIONES COMERCIALES Y FINANCIERAS.'])
        hoja.append(['Estado de la Ficha de Caracterización:', 'EN EJECUCIÓN'])
        hoja.append(['Fecha Inicio:', to_excel(date(2025, 9, 29))])
        hoja.append(['Fecha Fin:', '28/12/2026'])
        hoja.append(['Modalidad de Formación:', 'VIRTUAL'])
        hoja.append(['Regional:', '68 - REGIONAL SANTANDER'])
        hoja.append(['Centro de Formación:', 'CENTRO DE GESTIÓN AGROEMPRESARIAL DEL ORIENTE'])
        hoja.append([
            'Tipo de Documento',
            'Número de Documento',
            'Nombre',
            'Apellidos',
            'Estado',
            'Competencia',
            'Resultado de Aprendizaje',
            'Juicio de Evaluación',
            'Fecha y Hora del Juicio Evaluativo',
            'Funcionario que registro el juicio evaluativo',
        ])
        hoja.append([
            'CC',
            1000271292,
            'MARIO ANDRÉS',
            'CASTILLO MAESTRE',
            'EN FORMACION',
            '2 - RESULTADOS DE APRENDIZAJE ETAPA PRÁCTICA',
            '601427 - APLICAR EN LA RESOLUCIÓN DE PROBLEMAS',
            'APROBADO',
            '24/06/2026 08:15',
            'Instructor anterior',
        ])
        salida = BytesIO()
        libro.save(salida)
        salida.seek(0)
        return FileStorage(stream=salida, filename='Reporte de Juicios Evaluativos.xlsx')

    def test_reporte_consolida_aprendiz_y_juicio(self):
        resultado = importar_archivo(self._archivo(), self.ficha, self.a.id)
        db.session.commit()
        self.assertEqual(resultado['nuevos'], 1)
        self.assertEqual(resultado['juicios_nuevos'], 1)
        self.assertEqual(Aprendiz.query.count(), 1)
        self.assertEqual(JuicioEvaluativo.query.count(), 1)
        self.assertEqual(JuicioEvaluativoInstructor.query.count(), 1)

        resultado2 = importar_archivo(self._archivo(), self.ficha, self.b.id)
        db.session.commit()
        self.assertEqual(resultado2['juicios_nuevos'], 0)
        self.assertEqual(resultado2['juicios_repetidos'], 2)
        self.assertEqual(JuicioEvaluativo.query.count(), 1)
        self.assertEqual(FichaInstructor.query.count(), 2)
        self.assertEqual(JuicioEvaluativoInstructor.query.count(), 2)

    def test_crea_ficha_completa_desde_reporte(self):
        resultado = importar_archivo(
            self._reporte_para_crear(),
            ficha_actual=None,
            instructor_id=self.a.id,
            crear_ficha=True,
        )
        db.session.commit()

        ficha = resultado['ficha']
        aprendiz = Aprendiz.query.filter_by(ficha_id=ficha.id).one()
        self.assertTrue(resultado['ficha_creada'])
        self.assertEqual(ficha.codigo, '3336360')
        self.assertEqual(ficha.codigo_ficha, '3336360')
        self.assertEqual(ficha.codigo_programa, '133100')
        self.assertEqual(
            ficha.nombre_programa,
            'CONTABILIZACIÓN DE OPERACIONES COMERCIALES Y FINANCIERAS.',
        )
        self.assertEqual(ficha.fecha_inicio, date(2025, 9, 29))
        self.assertEqual(ficha.fecha_fin, date(2026, 12, 28))
        self.assertEqual(aprendiz.documento, '1000271292')
        self.assertEqual(aprendiz.estado, 'EN_FORMACION')
        self.assertEqual(ConfiguracionAlertas.query.filter_by(ficha_id=ficha.id).count(), 1)
        self.assertEqual(ConfiguracionRanking.query.filter_by(ficha_id=ficha.id).count(), 1)
        self.assertEqual(ConfiguracionAseo.query.filter_by(ficha_id=ficha.id).count(), 1)

    def test_ruta_crea_ficha_y_redirige_a_aprendices(self):
        cliente = self.app.test_client()
        cliente.post('/login', data={'correo': self.a.correo, 'password': 'x'})
        pagina = cliente.get('/instructor/fichas')
        self.assertIn(b'Crear ficha desde el reporte de juicios', pagina.data)
        self.assertIn(b'name="archivo"', pagina.data)
        reporte = self._reporte_para_crear()

        respuesta = cliente.post(
            '/instructor/fichas/importar-reporte',
            data={'archivo': (reporte.stream, reporte.filename)},
            content_type='multipart/form-data',
            follow_redirects=True,
        )

        ficha = Ficha.query.filter_by(codigo='3336360').one()
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.request.path, f'/instructor/fichas/{ficha.id}/aprendices')
        self.assertIn(b'creada autom', respuesta.data)
        self.assertEqual(Aprendiz.query.filter_by(ficha_id=ficha.id).count(), 1)

    def test_no_crea_ficha_desde_excel_sin_metadatos_oficiales(self):
        libro = openpyxl.Workbook()
        libro.active.append(['Documento', 'Nombre', 'Apellidos'])
        libro.active.append(['1001', 'Ana', 'Pérez'])
        salida = BytesIO()
        libro.save(salida)
        salida.seek(0)
        archivo = FileStorage(stream=salida, filename='lista.xlsx')

        with self.assertRaisesRegex(ErrorImportacion, 'Ficha de Caracterizaci'):
            importar_archivo(
                archivo,
                ficha_actual=None,
                instructor_id=self.a.id,
                crear_ficha=True,
            )
        self.assertEqual(Ficha.query.count(), 1)


if __name__ == '__main__':
    unittest.main()
