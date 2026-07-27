from collections import defaultdict
from datetime import datetime
import io

import openpyxl
from flask import Blueprint, flash, make_response, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl.styles import Alignment, Font, PatternFill

from app import db
from app.models.aprendiz import Aprendiz
from app.models.ficha import Ficha
from app.models.insignia import Insignia, InsigniaOtorgada
from app.models.ranking import PuntajeHistorico
from app.services.ranking import (
    asegurar_catalogo_insignias,
    actualizar_participacion_ficha,
    calcular_ranking,
    guardar_snapshot,
    obtener_configuracion,
)
from app.services.permisos import puede_gestionar_ficha


ranking_bp = Blueprint('ranking', __name__, template_folder='../templates/instructor')


def _ficha_autorizada(ficha_id):
    ficha = db.session.get(Ficha, ficha_id)
    if not ficha:
        return None
    if not puede_gestionar_ficha(ficha):
        return None
    return ficha


def _periodo():
    periodo = request.args.get('periodo', 'general')
    return periodo if periodo in ('general', 'semanal', 'mensual') else 'general'


def _evolucion(ficha_id, filas):
    historicos = PuntajeHistorico.query.filter_by(ficha_id=ficha_id).order_by(
        PuntajeHistorico.fecha_corte.asc()
    ).all()
    puntos = defaultdict(dict)
    for item in historicos:
        etiqueta = item.fecha_corte.strftime('%d/%m/%Y')
        puntos[item.aprendiz_id][etiqueta] = round(item.puntaje_total, 2)

    resultado = {}
    for fila in filas:
        aprendiz = fila['aprendiz']
        serie = puntos.get(aprendiz.id, {})
        if not serie:
            serie[datetime.utcnow().strftime('%d/%m/%Y')] = fila['puntaje_total']
        resultado[str(aprendiz.id)] = {
            'nombre': aprendiz.nombre_completo,
            'labels': list(serie.keys())[-24:],
            'valores': list(serie.values())[-24:],
        }
    return resultado


@ranking_bp.route('/fichas/<int:ficha_id>/ranking')
@login_required
def ranking(ficha_id):
    ficha = _ficha_autorizada(ficha_id)
    if not ficha:
        flash('Ficha no encontrada.', 'error')
        return redirect(url_for('instructor.fichas'))

    filas_generales, _ = actualizar_participacion_ficha(ficha_id)
    if _periodo() == 'general':
        filas = filas_generales
        config = obtener_configuracion(ficha_id)
    else:
        filas, config = calcular_ranking(ficha_id, periodo=_periodo())
    return render_template(
        'ranking.html',
        ficha=ficha,
        filas=filas,
        config=config,
        periodo=_periodo(),
        evolucion=_evolucion(ficha_id, filas),
    )


@ranking_bp.route('/fichas/<int:ficha_id>/ranking/lista')
@login_required
def ranking_lista(ficha_id):
    ficha = _ficha_autorizada(ficha_id)
    if not ficha:
        return '<div class="alert alert-error">Ficha no encontrada.</div>', 404
    filas, _ = calcular_ranking(ficha_id, periodo=_periodo())
    resp = make_response(render_template('instructor/_ranking_lista.html', ficha=ficha, filas=filas))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@ranking_bp.route('/fichas/<int:ficha_id>/ranking/configuracion', methods=['POST'])
@login_required
def configurar_ranking(ficha_id):
    ficha = _ficha_autorizada(ficha_id)
    if not ficha:
        flash('Ficha no encontrada.', 'error')
        return redirect(url_for('instructor.fichas'))

    config = obtener_configuracion(ficha_id, crear=True)
    try:
        peso_asistencia = float(request.form.get('peso_asistencia', 30))
        peso_evidencias = float(request.form.get('peso_evidencias', 40))
        peso_juicios = float(request.form.get('peso_juicios', 30))
        pesos = [peso_asistencia, peso_evidencias, peso_juicios]
        if abs(sum(pesos) - 100) > 0.01:
            raise ValueError('Los pesos deben sumar 100%.')
        if any(p < 0 for p in pesos):
            raise ValueError('Los pesos no pueden ser negativos.')

        config.peso_asistencia = peso_asistencia
        config.peso_evidencias = peso_evidencias
        config.peso_juicios = peso_juicios
        config.modo_visibilidad = (
            request.form.get('modo_visibilidad')
            if request.form.get('modo_visibilidad') in ('publico', 'privado')
            else 'privado'
        )
        config.modo_anonimo_parcial = 'modo_anonimo_parcial' in request.form
        config.periodo_corte = (
            request.form.get('periodo_corte')
            if request.form.get('periodo_corte') in ('mensual', 'trimestral', 'semestral')
            else 'trimestral'
        )
        config.bonus_entrega_anticipada = max(
            0, float(request.form.get('bonus_entrega_anticipada', 1))
        )
        config.horas_entrega_anticipada = max(
            0, int(request.form.get('horas_entrega_anticipada', 24))
        )
        config.bonus_racha_asistencia = max(
            0, float(request.form.get('bonus_racha_asistencia', 3))
        )
        config.semanas_racha = max(1, int(request.form.get('semanas_racha', 4)))
        config.bonus_calificacion_alta = max(
            0, float(request.form.get('bonus_calificacion_alta', 1))
        )
        config.umbral_calificacion_alta = max(
            0, float(request.form.get('umbral_calificacion_alta', 4))
        )
        config.penalizacion_falla_injustificada = max(
            0, float(request.form.get('penalizacion_falla_injustificada', 1))
        )
    except ValueError as error:
        db.session.rollback()
        flash(str(error) or 'Revisa los valores de configuración.', 'error')
        return redirect(url_for('ranking.ranking', ficha_id=ficha_id))

    db.session.commit()
    flash('Configuración del ranking actualizada.', 'success')
    return redirect(url_for('ranking.ranking', ficha_id=ficha_id))


@ranking_bp.route('/fichas/<int:ficha_id>/ranking/nuevo-corte', methods=['POST'])
@login_required
def nuevo_corte(ficha_id):
    ficha = _ficha_autorizada(ficha_id)
    if not ficha:
        flash('Ficha no encontrada.', 'error')
        return redirect(url_for('instructor.fichas'))

    ahora = datetime.utcnow()
    filas, config = calcular_ranking(ficha_id, periodo='general', ahora=ahora)
    guardar_snapshot(ficha_id, filas, ahora=ahora, tipo='corte')
    config.inicio_corte = ahora
    db.session.commit()
    flash(
        'Nuevo corte iniciado. El resultado anterior quedó guardado en el histórico.',
        'success',
    )
    return redirect(url_for('ranking.ranking', ficha_id=ficha_id))


@ranking_bp.route('/fichas/<int:ficha_id>/ranking/exportar')
@login_required
def exportar_ranking(ficha_id):
    ficha = _ficha_autorizada(ficha_id)
    if not ficha:
        flash('Ficha no encontrada.', 'error')
        return redirect(url_for('instructor.fichas'))

    periodo = _periodo()
    filas, _ = calcular_ranking(ficha_id, periodo=periodo)
    if request.args.get('formato') == 'pdf':
        return _ranking_pdf(ficha, filas, periodo)
    return _ranking_excel(ficha, filas, periodo)


def _ranking_excel(ficha, filas, periodo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ranking de participación'
    ws.append([f'Ranking de participación - Ficha {ficha.codigo}'])
    ws.append([
        'Herramienta motivacional. No corresponde a una nota oficial del SENA.'
    ])
    ws.append([])
    ws.append([
        'Posición',
        'Documento',
        'Aprendiz',
        'Asistencia %',
        'Evidencias %',
        'Juicios Aprob. %',
        'Bonificaciones',
        'Penalización',
        'Puntaje',
        'Tendencia',
    ])
    for fila in filas:
        ws.append([
            fila['posicion'],
            fila['aprendiz'].documento,
            fila['aprendiz'].nombre_completo,
            fila['porcentaje_asistencia'],
            fila['porcentaje_evidencias'],
            fila['porcentaje_aprobados'],
            fila['bonus'],
            fila['penalizacion'],
            fila['puntaje_total'],
            fila['tendencia'],
        ])

    verde = PatternFill('solid', fgColor='39A900')
    for celda in ws[4]:
        celda.fill = verde
        celda.font = Font(color='FFFFFF', bold=True)
        celda.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A5'
    for columna, ancho in {
        'A': 11, 'B': 18, 'C': 34, 'D': 15, 'E': 15,
        'F': 15, 'G': 16, 'H': 14, 'I': 12, 'J': 12,
    }.items():
        ws.column_dimensions[columna].width = ancho

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return send_file(
        salida,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'ranking_participacion_{ficha.codigo}_{periodo}.xlsx',
    )


def _ranking_pdf(ficha, filas, periodo):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from reportlab.lib.enums import TA_CENTER
    
    salida = io.BytesIO()
    doc = SimpleDocTemplate(salida, pagesize=landscape(letter), rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    estilos = getSampleStyleSheet()
    estilos['Title'].alignment = TA_CENTER
    estilos['Normal'].alignment = TA_CENTER
    estilos['Italic'].alignment = TA_CENTER
    elementos = [
        Paragraph(f'Ranking de participación - Ficha {ficha.codigo}', estilos['Title']),
        Paragraph(f'Periodo consultado: {periodo.capitalize()}', estilos['Normal']),
        Paragraph(
            'Herramienta motivacional; no corresponde a una nota oficial del SENA.',
            estilos['Italic'],
        ),
        Spacer(1, 12),
    ]
    datos = [[
        'Pos.', 'Aprendiz', 'Asist. %', 'Evid. %', 'Juicios %', 'Bonus', 'Penal.', 'Puntaje', 'Tend.'
    ]]
    for fila in filas:
        datos.append([
            fila['posicion'],
            fila['aprendiz'].nombre_completo,
            f"{fila['porcentaje_asistencia']:.1f}",
            f"{fila['porcentaje_evidencias']:.1f}",
            f"{fila['porcentaje_aprobados']:.1f}",
            f"{fila['bonus']:.1f}",
            f"{fila['penalizacion']:.1f}",
            f"{fila['puntaje_total']:.1f}",
            fila['tendencia'],
        ])
    ancho_util = doc.width
    tabla = Table(datos, colWidths=[ancho_util*0.05, ancho_util*0.34, ancho_util*0.08, ancho_util*0.08, ancho_util*0.08, ancho_util*0.08, ancho_util*0.08, ancho_util*0.08, ancho_util*0.08], repeatRows=1, hAlign='CENTER')
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39A900')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    salida.seek(0)
    return send_file(
        salida,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'ranking_participacion_{ficha.codigo}_{periodo}.pdf',
    )


@ranking_bp.route('/fichas/<int:ficha_id>/insignias')
@login_required
def insignias(ficha_id):
    ficha = _ficha_autorizada(ficha_id)
    if not ficha:
        flash('Ficha no encontrada.', 'error')
        return redirect(url_for('instructor.fichas'))

    asegurar_catalogo_insignias(ficha_id)
    db.session.commit()
    catalogo = Insignia.query.filter_by(ficha_id=ficha_id).order_by(
        Insignia.tipo, Insignia.nombre
    ).all()
    aprendices = Aprendiz.query_en_formacion(ficha_id).order_by(
        Aprendiz.apellidos, Aprendiz.nombre
    ).all()
    filtro_id = request.args.get('insignia_id', type=int)
    otorgamientos = InsigniaOtorgada.query.join(Insignia).filter(
        Insignia.ficha_id == ficha_id
    )
    if filtro_id:
        otorgamientos = otorgamientos.filter(InsigniaOtorgada.insignia_id == filtro_id)
    otorgamientos = otorgamientos.order_by(InsigniaOtorgada.fecha_obtencion.desc()).all()
    return render_template(
        'insignias.html',
        ficha=ficha,
        catalogo=catalogo,
        aprendices=aprendices,
        otorgamientos=otorgamientos,
        filtro_id=filtro_id,
    )


@ranking_bp.route('/fichas/<int:ficha_id>/insignias/crear', methods=['POST'])
@login_required
def crear_insignia(ficha_id):
    ficha = _ficha_autorizada(ficha_id)
    if not ficha:
        flash('Ficha no encontrada.', 'error')
        return redirect(url_for('instructor.fichas'))

    nombre = request.form.get('nombre', '').strip()
    descripcion = request.form.get('descripcion', '').strip()
    if not nombre or not descripcion:
        flash('El nombre y la descripción son obligatorios.', 'error')
        return redirect(url_for('ranking.insignias', ficha_id=ficha_id))

    insignia = Insignia(
        ficha_id=ficha_id,
        nombre=nombre,
        descripcion=descripcion,
        icono=(request.form.get('icono', '').strip() or '🏅')[:20],
        tipo='manual',
        condicion_json={'criterio': 'Otorgamiento según criterio pedagógico del instructor.'},
    )
    db.session.add(insignia)
    db.session.commit()
    flash(f'Insignia “{nombre}” creada.', 'success')
    return redirect(url_for('ranking.insignias', ficha_id=ficha_id))


@ranking_bp.route('/fichas/<int:ficha_id>/insignias/<int:insignia_id>/estado', methods=['POST'])
@login_required
def cambiar_estado_insignia(ficha_id, insignia_id):
    ficha = _ficha_autorizada(ficha_id)
    insignia = db.session.get(Insignia, insignia_id)
    if not ficha or not insignia or insignia.ficha_id != ficha_id:
        flash('Insignia no encontrada.', 'error')
        return redirect(url_for('instructor.fichas'))
    insignia.activa = not insignia.activa
    db.session.commit()
    estado = 'activada' if insignia.activa else 'desactivada'
    flash(f'Insignia {estado}. Los logros ya obtenidos se conservan.', 'success')
    return redirect(url_for('ranking.insignias', ficha_id=ficha_id))


@ranking_bp.route('/fichas/<int:ficha_id>/insignias/otorgar', methods=['POST'])
@login_required
def otorgar_insignia(ficha_id):
    ficha = _ficha_autorizada(ficha_id)
    aprendiz = db.session.get(Aprendiz, request.form.get('aprendiz_id', type=int))
    insignia = db.session.get(Insignia, request.form.get('insignia_id', type=int))
    if (
        not ficha
        or not aprendiz
        or not insignia
        or aprendiz.ficha_id != ficha_id
        or insignia.ficha_id != ficha_id
        or not insignia.activa
    ):
        flash('No fue posible otorgar la insignia seleccionada.', 'error')
        return redirect(url_for('ranking.insignias', ficha_id=ficha_id))

    existente = InsigniaOtorgada.query.filter_by(
        aprendiz_id=aprendiz.id, insignia_id=insignia.id
    ).first()
    if existente:
        flash('El aprendiz ya tiene esta insignia; los logros no se duplican.', 'info')
    else:
        db.session.add(InsigniaOtorgada(
            aprendiz_id=aprendiz.id,
            insignia_id=insignia.id,
            otorgada_por='instructor',
            instructor_id=current_user.id,
        ))
        db.session.commit()
        flash(f'¡Insignia “{insignia.nombre}” otorgada a {aprendiz.nombre_completo}!', 'success')
    return redirect(url_for('ranking.insignias', ficha_id=ficha_id))


@ranking_bp.route('/fichas/<int:ficha_id>/insignias/exportar')
@login_required
def exportar_insignias(ficha_id):
    ficha = _ficha_autorizada(ficha_id)
    if not ficha:
        flash('Ficha no encontrada.', 'error')
        return redirect(url_for('instructor.fichas'))

    aprendices = Aprendiz.query_en_formacion(ficha_id).order_by(
        Aprendiz.apellidos, Aprendiz.nombre
    ).all()
    otorgamientos = InsigniaOtorgada.query.join(Insignia).filter(
        Insignia.ficha_id == ficha_id
    ).order_by(InsigniaOtorgada.fecha_obtencion).all()
    por_aprendiz = defaultdict(list)
    for item in otorgamientos:
        por_aprendiz[item.aprendiz_id].append(item)
    if request.args.get('formato') == 'pdf':
        return _insignias_pdf(ficha, aprendices, por_aprendiz)
    return _insignias_excel(ficha, aprendices, por_aprendiz)


def _insignias_excel(ficha, aprendices, por_aprendiz):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reconocimientos'
    ws.append([f'Reconocimientos de participación - Ficha {ficha.codigo}'])
    ws.append(['Estos reconocimientos no hacen parte de la evaluación oficial.'])
    ws.append([])
    ws.append(['Documento', 'Aprendiz', 'Insignia', 'Fecha', 'Otorgada por'])
    for aprendiz in aprendices:
        items = por_aprendiz.get(aprendiz.id) or [None]
        for item in items:
            ws.append([
                aprendiz.documento,
                aprendiz.nombre_completo,
                item.insignia.nombre if item else 'Sin insignias todavía',
                item.fecha_obtencion.strftime('%d/%m/%Y') if item else '',
                item.otorgada_por if item else '',
            ])
    for celda in ws[4]:
        celda.fill = PatternFill('solid', fgColor='39A900')
        celda.font = Font(color='FFFFFF', bold=True)
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return send_file(
        salida,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'insignias_participacion_{ficha.codigo}.xlsx',
    )


def _insignias_pdf(ficha, aprendices, por_aprendiz):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from reportlab.lib.enums import TA_CENTER
    
    salida = io.BytesIO()
    doc = SimpleDocTemplate(salida, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    estilos = getSampleStyleSheet()
    estilos['Title'].alignment = TA_CENTER
    estilos['Italic'].alignment = TA_CENTER
    elementos = [
        Paragraph(f'Reconocimientos de participación - Ficha {ficha.codigo}', estilos['Title']),
        Paragraph('No hacen parte de la evaluación oficial del aprendiz.', estilos['Italic']),
        Spacer(1, 12),
    ]
    datos = [['Aprendiz', 'Insignias obtenidas']]
    for aprendiz in aprendices:
        nombres = ', '.join(
            item.insignia.nombre for item in por_aprendiz.get(aprendiz.id, [])
        ) or 'Sin insignias todavía'
        datos.append([aprendiz.nombre_completo, nombres])
    ancho_util = doc.width
    tabla = Table(datos, colWidths=[ancho_util*0.4, ancho_util*0.6], repeatRows=1, hAlign='CENTER')
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#39A900')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    salida.seek(0)
    return send_file(
        salida,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'insignias_participacion_{ficha.codigo}.pdf',
    )
